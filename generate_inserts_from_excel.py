#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script สำหรับสร้าง SQL INSERT statements จากไฟล์ Excel
และเพิ่มคำสั่งลบข้อมูลก่อน INSERT
"""

import openpyxl
import re
from datetime import datetime
from collections import defaultdict

# ไฟล์ Excel ที่ต้องการอ่าน
EXCEL_FILE = 'ครุภัณฑ์_รวม_2568.xlsx'
OUTPUT_FILE = 'insert_statements_with_reset.sql'

def parse_price(price_str):
    """แปลงราคาจาก string เป็น float"""
    if not price_str:
        return 0.00
    
    # ลบเครื่องหมาย comma และ space
    price_str = str(price_str).replace(',', '').replace(' ', '').strip()
    
    try:
        price = float(price_str)
        # ตรวจสอบว่าเกิน limit ของ NUMERIC(12, 2) หรือไม่ (9,999,999,999.99)
        max_price = 9999999999.99
        if price > max_price:
            print(f"⚠️  Warning: Price {price} exceeds maximum, setting to 0.00")
            return 0.00
        return round(price, 2)
    except (ValueError, TypeError):
        return 0.00

def extract_date_from_code(code):
    """ดึงวันที่ซื้อจากรหัสทรัพย์สิน (รูปแบบ: PREFIX-DD-MM-YYYY)"""
    if not code or not isinstance(code, str):
        return None
    
    # รูปแบบ: A004-09-04-2557 -> ดึง 09-04-2557
    match = re.match(r'^[A-Z]+\d+-(\d{1,2})-(\d{1,2})-(\d{4})$', code)
    if not match:
        return None
    
    try:
        day = int(match.group(1))
        month = int(match.group(2))
        year = int(match.group(3))
        
        # แปลง พ.ศ. เป็น ค.ศ. (ถ้า > 2500)
        if year > 2500:
            year = year - 543
        
        # ตรวจสอบความถูกต้อง
        if year < 1900 or year > 2100:
            return None
        if month < 1 or month > 12:
            return None
        if day < 1 or day > 31:
            return None
        
        # สร้างวันที่ (format: YYYY-MM-DD)
        return f"{year}-{month:02d}-{day:02d}"
    except (ValueError, IndexError):
        return None

def convert_thai_date(date_str):
    """แปลงวันที่จากรูปแบบไทยเป็น YYYY-MM-DD"""
    if not date_str:
        return None
    
    date_str = str(date_str).strip()
    
    # ถ้าเป็นรูปแบบ DD/MM/YYYY หรือ DD-MM-YYYY
    match = re.match(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        year = int(year)
        if year > 2500:
            year = year - 543
        try:
            return f"{year}-{int(month):02d}-{int(day):02d}"
        except:
            return None
    
    # ถ้าเป็นรูปแบบ YYYY-MM-DD อยู่แล้ว
    match = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
    if match:
        year, month, day = match.groups()
        year = int(year)
        if year > 2500:
            year = year - 543
        try:
            return f"{year}-{int(month):02d}-{int(day):02d}"
        except:
            return None
    
    return None

def escape_sql_string(value):
    """Escape string สำหรับ SQL"""
    if value is None:
        return 'NULL'
    if isinstance(value, (int, float)):
        return str(value)
    # Escape single quotes
    escaped = str(value).replace("'", "''")
    return f"'{escaped}'"

def get_category_icon(category_name):
    """กำหนด icon สำหรับหมวดหมู่"""
    icon_map = {
        'คอมพิวเตอร์': 'Monitor',
        'Printer': 'Printer',
        'กล้องถ่ายรูป': 'Camera',
        'ตู้เก็บเอกสาร': 'Archive',
        'เครื่องเสียง': 'Speaker',
        'เฟอร์นิเจอร์': 'Sofa',
        'เครื่องใช้ไฟฟ้า': 'Fan',
        'อุปกรณ์สำนักงาน': 'FileText',
        'ยานพาหนะ': 'Car',
        'อุปกรณ์อิเล็กทรอนิกส์': 'HardDrive',
    }
    
    for key, icon in icon_map.items():
        if key in category_name:
            return icon
    
    return 'Package'  # default

def main():
    print(f"📖 กำลังอ่านไฟล์ Excel: {EXCEL_FILE}")
    
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except FileNotFoundError:
        print(f"❌ ไม่พบไฟล์: {EXCEL_FILE}")
        return
    except Exception as e:
        print(f"❌ เกิดข้อผิดพลาดในการอ่านไฟล์: {e}")
        return
    
    # เก็บหมวดหมู่ที่พบ
    categories = {}
    assets = []
    seen_codes = set()
    used_prefixes = set()  # เก็บ prefix ที่ใช้แล้ว
    
    # อ่านทุก sheet
    for sheet_name in workbook.sheetnames:
        print(f"📄 กำลังประมวลผล sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # หา header row (แถวแรกที่มีข้อมูล)
        header_row = 1
        headers = {}
        
        # อ่าน header row
        for col_idx, cell in enumerate(sheet[header_row], 1):
            if cell.value:
                header_text = str(cell.value).strip()
                headers[header_text] = col_idx
        
        print(f"  📋 Headers: {list(headers.keys())}")
        
        # อ่านข้อมูล (เริ่มจากแถวที่ 2)
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            
            # อ่านค่าจากคอลัมน์ตาม header
            category_code = str(row[headers.get('รหัสหมวด', 1) - 1].value).strip() if row[headers.get('รหัสหมวด', 1) - 1].value else ''
            category_name = str(row[headers.get('ชื่อหมวด', 2) - 1].value).strip() if row[headers.get('ชื่อหมวด', 2) - 1].value else ''
            code = str(row[headers.get('เลขทะเบียน', 4) - 1].value).strip() if row[headers.get('เลขทะเบียน', 4) - 1].value else ''
            purchase_date_str = str(row[headers.get('วันที่ซื้อ', 5) - 1].value).strip() if row[headers.get('วันที่ซื้อ', 5) - 1].value else ''
            brand = str(row[headers.get('ยี่ห้อ', 6) - 1].value).strip() if row[headers.get('ยี่ห้อ', 6) - 1].value else ''
            color = str(row[headers.get('สี', 7) - 1].value).strip() if row[headers.get('สี', 7) - 1].value else ''
            model = str(row[headers.get('รุ่น', 8) - 1].value).strip() if row[headers.get('รุ่น', 8) - 1].value else ''
            serial = str(row[headers.get('Serial No.', 9) - 1].value).strip() if row[headers.get('Serial No.', 9) - 1].value else ''
            price = row[headers.get('ราคาทุน', 10) - 1].value if row[headers.get('ราคาทุน', 10) - 1].value else None
            
            # อ่านคอลัมน์เพิ่มเติมถ้ามี
            location = None
            status = 'Normal'
            
            # หาคอลัมน์สถานที่ (ลองหลายชื่อ)
            location_col = None
            for loc_key in ['สถานที่ใช้งานปัจจุบัน', 'สถานที่ตั้ง', 'สถานที่', 'ที่ตั้ง', 'location']:
                if loc_key in headers:
                    location_col = headers[loc_key]
                    break
            
            if location_col:
                location = str(row[location_col - 1].value).strip() if row[location_col - 1].value else None
                if location and location.lower() in ['', 'null', 'none', '-']:
                    location = None
            
            # หาคอลัมน์สถานะ
            status_col = None
            for status_key in ['สถานะ', 'status', 'สภานะ']:
                if status_key in headers:
                    status_col = headers[status_key]
                    break
            
            if status_col:
                status = str(row[status_col - 1].value).strip() if row[status_col - 1].value else 'Normal'
            
            # ข้ามแถวว่างหรือแถวสรุป
            if not code or code.startswith('รวม') or code.startswith('รวมหมวด') or code == 'รหัสหมวด':
                continue
            
            # ตรวจสอบว่า code เป็นวันที่หรือไม่
            if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{4}$', code) or re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', code):
                continue
            
            # ตรวจสอบ code ซ้ำ
            if code in seen_codes:
                print(f"  ⚠️  ข้าม code ซ้ำ: {code} (แถว {row_idx})")
                continue
            seen_codes.add(code)
            
            # ดึงวันที่ซื้อ
            purchase_date = convert_thai_date(purchase_date_str)
            
            # ถ้าไม่มีวันที่ ให้ดึงจาก code
            if not purchase_date:
                purchase_date = extract_date_from_code(code)
            
            # ตรวจสอบหมวดหมู่
            if not category_name:
                category_name = 'อื่นๆ'
            
            # สร้างชื่อทรัพย์สินจาก brand + model
            name_parts = []
            if brand:
                name_parts.append(brand)
            if model:
                name_parts.append(model)
            name = ' '.join(name_parts) if name_parts else 'ไม่ระบุ'
            
            # เพิ่มหมวดหมู่ถ้ายังไม่มี
            if category_name not in categories:
                # กำหนด prefix โดยตรวจสอบว่าไม่ซ้ำ
                prefix = category_code if category_code else category_name[0].upper()
                
                # ถ้า prefix ซ้ำ ให้หาค่าที่ไม่ซ้ำ
                original_prefix = prefix
                counter = 1
                while prefix in used_prefixes:
                    # ลองใช้ prefix + ตัวเลข
                    prefix = f"{original_prefix}{counter}"
                    counter += 1
                    # ถ้าเกิน 9 ให้ใช้ตัวอักษรถัดไป
                    if counter > 9:
                        # ใช้ตัวอักษรถัดไป
                        if len(original_prefix) == 1 and original_prefix.isalpha():
                            next_char = chr(ord(original_prefix) + 1)
                            if next_char <= 'Z':
                                prefix = next_char
                            else:
                                prefix = f"{original_prefix}1"
                        else:
                            prefix = f"{original_prefix}1"
                        break
                
                used_prefixes.add(prefix)
                icon_name = get_category_icon(category_name)
                categories[category_name] = {
                    'name': category_name,
                    'prefix': prefix,
                    'useful_life': 5,  # default
                    'icon_name': icon_name
                }
            
            # แปลงราคา
            price_value = parse_price(price)
            
            # เพิ่มทรัพย์สิน
            assets.append({
                'code': code,
                'name': name,
                'brand': brand or None,
                'serial': serial or None,
                'price': price_value,
                'location': location,
                'status': status,
                'purchase_date': purchase_date,
                'category': category_name,
                'useful_life': categories[category_name]['useful_life'],
                'color': color or None,
                'image': None
            })
    
    print(f"\n✅ พบหมวดหมู่: {len(categories)} หมวด")
    print(f"✅ พบทรัพย์สิน: {len(assets)} รายการ")
    
    # สร้าง SQL file
    print(f"\n📝 กำลังสร้างไฟล์ SQL: {OUTPUT_FILE}")
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        # เขียน header
        f.write("-- ================================================================================\n")
        f.write("-- SQL Script: Reset และ Import ข้อมูลจาก Excel\n")
        f.write(f"-- สร้างเมื่อ: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("-- ================================================================================\n\n")
        
        # คำสั่งลบข้อมูล
        f.write("-- ================================================================================\n")
        f.write("-- PART 1: ลบข้อมูลเดิม\n")
        f.write("-- ================================================================================\n\n")
        
        f.write("-- ลบ audit logs ทั้งหมด\n")
        f.write("DELETE FROM audit_logs;\n\n")
        
        f.write("-- ลบข้อมูลทั้งหมดในตาราง assets\n")
        f.write("TRUNCATE TABLE assets RESTART IDENTITY CASCADE;\n\n")
        
        f.write("-- ลบหมวดหมู่ทั้งหมด (ถ้าต้องการเริ่มใหม่)\n")
        f.write("-- TRUNCATE TABLE categories RESTART IDENTITY CASCADE;\n\n")
        
        # INSERT categories
        f.write("-- ================================================================================\n")
        f.write("-- PART 2: INSERT Categories\n")
        f.write("-- ================================================================================\n\n")
        
        for cat_name, cat_data in sorted(categories.items()):
            # ใช้ DO NOTHING ถ้า prefix ซ้ำ (จะไม่ error และไม่สร้างใหม่)
            # ใช้ DO UPDATE ถ้า name ซ้ำ (อัพเดทข้อมูล)
            f.write(f"-- หมวดหมู่: {cat_data['name']} (prefix: {cat_data['prefix']})\n")
            f.write(f"INSERT INTO categories (name, prefix, useful_life, icon_name)\n")
            f.write(f"VALUES ({escape_sql_string(cat_data['name'])}, {escape_sql_string(cat_data['prefix'])}, {cat_data['useful_life']}, {escape_sql_string(cat_data['icon_name'])})\n")
            f.write(f"ON CONFLICT (name) DO UPDATE SET\n")
            f.write(f"    useful_life = EXCLUDED.useful_life,\n")
            f.write(f"    icon_name = EXCLUDED.icon_name;\n\n")
        
        # INSERT assets
        f.write("-- ================================================================================\n")
        f.write("-- PART 3: INSERT Assets\n")
        f.write("-- ================================================================================\n\n")
        
        for asset in assets:
            f.write(f"INSERT INTO assets (code, name, brand, serial, price, location, status, purchase_date, category, useful_life, color, image, is_sticker_printed)\n")
            f.write(f"VALUES (\n")
            f.write(f"    {escape_sql_string(asset['code'])},\n")
            f.write(f"    {escape_sql_string(asset['name'])},\n")
            f.write(f"    {escape_sql_string(asset['brand'])},\n")
            f.write(f"    {escape_sql_string(asset['serial'])},\n")
            f.write(f"    {asset['price']},\n")
            f.write(f"    {escape_sql_string(asset['location'])},\n")
            f.write(f"    {escape_sql_string(asset['status'])},\n")
            f.write(f"    {escape_sql_string(asset['purchase_date'])},\n")
            f.write(f"    {escape_sql_string(asset['category'])},\n")
            f.write(f"    {asset['useful_life']},\n")
            f.write(f"    {escape_sql_string(asset['color'])},\n")
            f.write(f"    {escape_sql_string(asset['image'])},\n")
            f.write(f"    false\n")
            f.write(f")\n")
            f.write(f"ON CONFLICT (code) DO UPDATE SET\n")
            f.write(f"    name = EXCLUDED.name,\n")
            f.write(f"    brand = EXCLUDED.brand,\n")
            f.write(f"    serial = EXCLUDED.serial,\n")
            f.write(f"    price = EXCLUDED.price,\n")
            f.write(f"    location = EXCLUDED.location,\n")
            f.write(f"    status = EXCLUDED.status,\n")
            f.write(f"    purchase_date = EXCLUDED.purchase_date,\n")
            f.write(f"    category = EXCLUDED.category,\n")
            f.write(f"    useful_life = EXCLUDED.useful_life,\n")
            f.write(f"    color = EXCLUDED.color,\n")
            f.write(f"    image = EXCLUDED.image;\n\n")
        
        # Summary
        f.write("-- ================================================================================\n")
        f.write("-- Summary\n")
        f.write("-- ================================================================================\n")
        f.write(f"-- Categories: {len(categories)}\n")
        f.write(f"-- Assets: {len(assets)}\n")
        f.write(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    print(f"✅ สร้างไฟล์ SQL สำเร็จ: {OUTPUT_FILE}")
    print(f"\n📊 สรุป:")
    print(f"   - หมวดหมู่: {len(categories)} หมวด")
    print(f"   - ทรัพย์สิน: {len(assets)} รายการ")
    print(f"\n💡 วิธีใช้งาน:")
    print(f"   1. เปิดไฟล์ {OUTPUT_FILE}")
    print(f"   2. คัดลอกเนื้อหาทั้งหมด")
    print(f"   3. วางใน Supabase SQL Editor")
    print(f"   4. รัน SQL script")

if __name__ == '__main__':
    main()

