#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
สคริปต์สำหรับสร้างคำสั่ง INSERT จากไฟล์ Excel
"""

import openpyxl
import re
from datetime import datetime

def clean_value(value):
    """ทำความสะอาดค่า"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(value).strip() if value else None
    return str(value).strip() if str(value).strip() else None

def parse_date(date_value):
    """แปลงวันที่ (รองรับทั้ง พ.ศ. และ ค.ศ.)"""
    if date_value is None:
        return None
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    if isinstance(date_value, str):
        # ลองแปลงรูปแบบวันที่ต่างๆ
        try:
            # รูปแบบ YYYY-MM-DD (ค.ศ.)
            if re.match(r'\d{4}-\d{2}-\d{2}', date_value):
                year = int(date_value.split('-')[0])
                # ถ้าเป็น พ.ศ. (มากกว่า 2500) แปลงเป็น ค.ศ.
                if year > 2500:
                    year = year - 543
                    return f"{year}-{date_value.split('-')[1]}-{date_value.split('-')[2]}"
                return date_value
            # รูปแบบ DD/MM/YYYY
            if re.match(r'\d{1,2}/\d{1,2}/\d{4}', date_value):
                parts = date_value.split('/')
                year = int(parts[2])
                # ถ้าเป็น พ.ศ. (มากกว่า 2500) แปลงเป็น ค.ศ.
                if year > 2500:
                    year = year - 543
                return f"{year}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        except:
            pass
    return None

def parse_price(price_value):
    """แปลงราคา"""
    if price_value is None:
        return 0.00
    if isinstance(price_value, (int, float)):
        return float(price_value)
    if isinstance(price_value, str):
        # ลบเครื่องหมายคอมมาและสัญลักษณ์อื่นๆ
        cleaned = re.sub(r'[^\d.]', '', price_value)
        try:
            return float(cleaned) if cleaned else 0.00
        except:
            return 0.00
    return 0.00

def escape_sql_string(value):
    """Escape string สำหรับ SQL"""
    if value is None:
        return 'NULL'
    return "'" + str(value).replace("'", "''").replace("\\", "\\\\") + "'"

def get_category_name(sheet_name):
    """ดึงชื่อหมวดหมู่จากชื่อ sheet"""
    # เช่น "A (คอมพิวเตอร์)" -> "คอมพิวเตอร์"
    match = re.search(r'\(([^)]+)\)', sheet_name)
    if match:
        return match.group(1)
    # ถ้าไม่มีวงเล็บ ใช้ชื่อ sheet ทั้งหมด
    return sheet_name.strip()

def get_prefix(sheet_name):
    """ดึง prefix จากชื่อ sheet"""
    match = re.match(r'^([A-Za-z])', sheet_name)
    return match.group(1).upper() if match else None

# เปิดไฟล์ Excel
import sys
output_file = 'insert_statements.sql'
if len(sys.argv) > 1:
    output_file = sys.argv[1]

wb = openpyxl.load_workbook('ครุภัณฑ์ 2568_kt.xlsx', data_only=True)

# หา sheet ที่ขึ้นต้นด้วยอักษรอังกฤษเท่านั้น (กรอง sheet ที่ไม่เกี่ยวข้อง)
english_prefix_sheets = []
for sheet_name in wb.sheetnames:
    if re.match(r'^[A-Za-z]', sheet_name):
        english_prefix_sheets.append(sheet_name)

print(f"📋 พบ sheet ที่ขึ้นต้นด้วยอักษรอังกฤษ: {len(english_prefix_sheets)} แผ่น")
print(f"   {', '.join(english_prefix_sheets[:5])}{'...' if len(english_prefix_sheets) > 5 else ''}")

# เก็บหมวดหมู่
categories = {}
assets = []

# อ่านข้อมูลจากแต่ละ sheet
for sheet_name in english_prefix_sheets:
    prefix = get_prefix(sheet_name)
    category_name = get_category_name(sheet_name)
    
    if category_name not in categories:
        categories[category_name] = prefix
    
    sheet = wb[sheet_name]
    
    # หาแถว header (แถวที่ 5)
    data_start_row = 6
    
    row_count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        # ตรวจสอบว่าแถวนี้มีข้อมูลหรือไม่
        if not any(row[2:14] if len(row) > 14 else row[2:]):  # ตรวจสอบคอลัมน์ C-N
            continue
        
        # อ่านข้อมูลตามคอลัมน์
        # คอลัมน์: A=ลำดับ, C=เลขทะเบียน, D=วันที่ซื้อ, E=ยี่ห้อ, F=สี, G=รุ่น, H=Serial No., 
        #          I=ราคาทุน, J=สถานที่ซื้อ, K=สถานที่ใช้งานปัจจุบัน, L=ผู้ตรวจนับ, M=เลขครุภัณฑ์, N=แปะสติกเกอร์
        
        asset_code = clean_value(row[2]) if len(row) > 2 else None  # เลขทะเบียน (C)
        if not asset_code:
            continue
        
        # ดึงวันที่ซื้อจากคอลัมน์ D ก่อน
        purchase_date = parse_date(row[3]) if len(row) > 3 else None  # วันที่ซื้อ (D)
        
        # ถ้าไม่มีวันที่ซื้อ ให้ลองดึงจากรหัสทรัพย์สิน (รูปแบบ: A004-09-04-2557)
        if not purchase_date and asset_code:
            date_match = re.match(r'.*-(\d{1,2})-(\d{1,2})-(\d{4})$', str(asset_code))
            if date_match:
                try:
                    day = int(date_match.group(1))
                    month = int(date_match.group(2))
                    year = int(date_match.group(3))
                    # แปลง พ.ศ. เป็น ค.ศ. (ถ้า > 2500)
                    if year > 2500:
                        year = year - 543
                    # ตรวจสอบความถูกต้อง
                    if 1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                        purchase_date = f"{year}-{month:02d}-{day:02d}"
                except:
                    pass
        brand = clean_value(row[4]) if len(row) > 4 else None  # ยี่ห้อ (E)
        color = clean_value(row[5]) if len(row) > 5 else None  # สี (F)
        model = clean_value(row[6]) if len(row) > 6 else None  # รุ่น (G)
        serial = clean_value(row[7]) if len(row) > 7 else None  # Serial No. (H)
        price = parse_price(row[8]) if len(row) > 8 else None  # ราคาทุน (I)
        purchase_location = clean_value(row[9]) if len(row) > 9 else None  # สถานที่ซื้อ (J)
        current_location = clean_value(row[10]) if len(row) > 10 else None  # สถานที่ใช้งานปัจจุบัน (K)
        checker = clean_value(row[11]) if len(row) > 11 else None  # ผู้ตรวจนับ (L)
        asset_number = clean_value(row[12]) if len(row) > 12 else None  # เลขครุภัณฑ์ (M)
        sticker_printed = clean_value(row[13]) if len(row) > 13 else None  # แปะสติกเกอร์ (N)
        
        # สร้างชื่อทรัพย์สิน (ยี่ห้อ + รุ่น + สี)
        name_parts = []
        if brand:
            name_parts.append(str(brand))
        if model:
            name_parts.append(str(model))
        if color:
            name_parts.append(str(color))
        asset_name = ' '.join(name_parts) if name_parts else asset_code
        
        # ใช้สถานที่ใช้งานปัจจุบัน หรือสถานที่ซื้อ
        location = current_location or purchase_location or None
        
        # สถานะ (default Normal)
        status = 'Normal'
        
        # อายุการใช้งาน (default 5)
        useful_life = 5
        
        # ตรวจสอบว่ารหัสเป็นรูปแบบที่ถูกต้อง (ไม่ใช่แค่วันที่)
        # รูปแบบที่ถูกต้อง: มี prefix (A-Z) + ตัวเลข + - + วันที่ หรือรูปแบบอื่นๆ
        is_valid_code = True
        if asset_code:
            # ตรวจสอบว่าไม่ใช่แค่วันที่ (รูปแบบ DD/MM/YYYY หรือ YYYY-MM-DD)
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', str(asset_code)) or \
               re.match(r'^\d{4}-\d{2}-\d{2}', str(asset_code)):
                is_valid_code = False
                print(f"⚠️  ข้าม: รหัส '{asset_code}' เป็นวันที่ ไม่ใช่รหัสทรัพย์สิน (แถว {row_idx})")
        
        if is_valid_code:
            assets.append({
                'code': asset_code,
                'name': asset_name,
                'brand': brand,
                'color': color,
                'serial': serial,
                'price': price,
                'location': location,
                'status': status,
                'purchase_date': purchase_date,
                'category': category_name,
                'useful_life': useful_life,
                'prefix': prefix
            })
            
            row_count += 1

# สร้างคำสั่ง INSERT สำหรับหมวดหมู่
output_lines = []
output_lines.append("-- ================================================================================")
output_lines.append("-- INSERT statements for Categories (หมวดครุภัณฑ์)")
output_lines.append("-- ================================================================================")
output_lines.append("")

category_inserts = []
for category_name, prefix in sorted(categories.items()):
    category_inserts.append(f"INSERT INTO categories (name) VALUES ({escape_sql_string(category_name)});")

for stmt in category_inserts:
    output_lines.append(stmt)

output_lines.append("")
output_lines.append(f"-- Total: {len(category_inserts)} categories")
output_lines.append("")
output_lines.append("")

# สร้างคำสั่ง INSERT สำหรับครุภัณฑ์
output_lines.append("-- ================================================================================")
output_lines.append("-- INSERT statements for Assets (ข้อมูลครุภัณฑ์)")
output_lines.append("-- ================================================================================")
output_lines.append("")

asset_inserts = []
seen_codes = {}  # เก็บรหัสที่เห็นแล้วเพื่อตรวจสอบซ้ำ

for asset in assets:
    code = asset['code']
    
    # ตรวจสอบรหัสซ้ำ
    if code in seen_codes:
        seen_codes[code] += 1
        print(f"⚠️  ข้าม: รหัส '{code}' ซ้ำ (ครั้งที่ {seen_codes[code]})")
        continue
    else:
        seen_codes[code] = 1
    
    code_escaped = escape_sql_string(code)
    name = escape_sql_string(asset['name'])
    brand = escape_sql_string(asset['brand'])
    color = escape_sql_string(asset['color']) if asset.get('color') else 'NULL'
    serial = escape_sql_string(asset['serial']) if asset.get('serial') else 'NULL'
    price = f"{asset['price']:.2f}"
    location = escape_sql_string(asset['location']) if asset.get('location') else 'NULL'
    status = escape_sql_string(asset['status'])
    purchase_date = escape_sql_string(asset['purchase_date']) if asset.get('purchase_date') else 'NULL'
    category = escape_sql_string(asset['category']) if asset.get('category') else 'NULL'
    useful_life = asset['useful_life']
    
    # ใช้ ON CONFLICT DO UPDATE เพื่อจัดการกรณีรหัสซ้ำ
    stmt = f"INSERT INTO assets (code, name, brand, color, serial, price, location, status, purchase_date, category, useful_life) VALUES ({code_escaped}, {name}, {brand}, {color}, {serial}, {price}, {location}, {status}, {purchase_date}, {category}, {useful_life}) ON CONFLICT (code) DO UPDATE SET name = EXCLUDED.name, brand = EXCLUDED.brand, color = EXCLUDED.color, serial = EXCLUDED.serial, price = EXCLUDED.price, location = EXCLUDED.location, status = EXCLUDED.status, purchase_date = EXCLUDED.purchase_date, category = EXCLUDED.category, useful_life = EXCLUDED.useful_life;"
    asset_inserts.append(stmt)

for stmt in asset_inserts:
    output_lines.append(stmt)

output_lines.append("")
output_lines.append(f"-- Total: {len(asset_inserts)} assets")
output_lines.append("")
output_lines.append("-- ================================================================================")
output_lines.append("-- Summary")
output_lines.append("-- ================================================================================")
output_lines.append(f"-- Categories: {len(categories)}")
output_lines.append(f"-- Assets: {len(assets)}")
output_lines.append("-- ================================================================================")

# เขียนไฟล์
with open(output_file, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))

print(f"✅ สร้างไฟล์ {output_file} สำเร็จ!")
print(f"   - หมวดหมู่: {len(categories)} หมวด")
print(f"   - ครุภัณฑ์: {len(assets)} รายการ")

